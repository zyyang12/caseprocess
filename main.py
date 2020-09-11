from openpyxl import Workbook
from openpyxl import load_workbook
from collections import OrderedDict
import os
import json
from parseconf import ParseConf
from directory import Directory

def readconfig():
    parseconf = ParseConf("config.ini")
    directory1 = {}
    rule1 = {}
    directory2 = {}
    rule2 = {}
    function = {}
    directory1["input"] = parseconf.parseStr("directory1", "input")
    directory1["result"] = parseconf.parseStr("directory1", "result")
    function["function"] = parseconf.parseStr("function", "function")
    rule1["condition"] = parseconf.parseList("rule1", "condition")
    rule1["flag"] = parseconf.parseStr("rule1", "flag")
    rule1["multiflag"] = parseconf.parseStr("rule1", "multiflag")
    rule1["nullflag"] = parseconf.parseStr("rule1", "nullflag")
    rule1["singlefileflag"] = parseconf.parseStr("rule1", "singlefileflag")
    directory2["input"] = parseconf.parseStr("directory2", "input")
    directory2["output"] = parseconf.parseStr("directory2", "output")
    rule2["key"] = parseconf.parseStr("rule2", "key")
    configdict ={"function":function,"directory1":directory1, "rule1":rule1, "directory2":directory2, "rule2":rule2}
    return configdict

def getFilelist(path):
    filelist = []
    directory = path.strip()
    templist = directory.split(".")
    if (templist[-1] == "xlsx"):
        filelist.append(directory)
    else:
        files = Directory.getFiles(directory)
        for file in files:
            templist = file.split(".")
            if (templist[-1] == "xlsx"):
                filelist.append(file)
    return filelist

def read_excel(filename):
    '''
    提取excel文件数据
    '''
    datasdict_list = []
    dataslist_list = []
    title_list = []
    wb = load_workbook(filename)
    for sheetname in wb.sheetnames:
        #datas[sheetname] = []
        top_flag = True
        title_list = []
        datadict_list = []
        datalist_list = []
        for row in wb[sheetname].rows:
            #data_dict = OrderedDict()
            data_list = []
            if top_flag:
                for cell in row:
                    title_list.append(cell.value)
                top_flag = False
            else:
                data_dict = OrderedDict()
                data_list = []
                for index in range(0, len(row)):
                    data_dict [title_list[index]] = row[index].value
                    data_list.append(row[index].value)
                datadict_list.append(data_dict)
                datalist_list.append(data_list)
                #print(datadict_list)
        datasdict_list.extend(datadict_list)
        dataslist_list.extend(datalist_list)
    wb.close()
    return datasdict_list, dataslist_list, title_list

def write_excel(filename, datalist,titlelist, flag):
    '''
    将合并后的数据写入excel文件
    '''
    #print(datalist[0].keys())
    wb = Workbook()
    ws = wb.active
    #titlelist = []
    if flag == "1":
        for i in range(len(datalist)-1, -1,-1):
            if datalist[i][-1] == 1:
                del datalist[i]
            else:
                del datalist[i][-1]
    elif flag == "0":
        titlelist.append("flag")
    '''
    if (len(datalist)>0):
        titlelist = list(datalist[0].keys())
    '''
    print(datalist)
    ws.append(titlelist)
    for i in range(len(datalist)):
        ws.append(datalist[i])
    wb.save(filename)
    wb.close()
    return 0

def de_weight(filelist, conditions, flag, multiflag, nullflag,resultpath):
    count = 0
    filedict = {}
    filetitledict = {}
    datalist = []
    databaklist = []
    uidlist = []
    conditionlist = []
    print(filelist)
    for file in filelist:
        filelist = file.split("\\")
        filename = filelist[-1]
        uidlist = []
        datas, datasbak, titlelist = read_excel(file)
        filetitledict[filename] = titlelist
        for i in range(len(datas)):
            datas[i]["index"] = i + count
            datas[i]["delindex"] = [i+count]
            #print(datas[i])
            tmp = ""
            #print(conditions)
            for condition in conditions:
                tmp = tmp + str(datas[i][condition])
            #print(tmp)
                #将要比较的字段拼接在一起
            if (nullflag == "0"):
                # 去重参数为空时（第一个参数），不参与去重
                if not datas[i][conditions[0]]:
                    #print(datas[i][conditions[0]])
                    #print(conditions[0])
                    tmp = tmp + str(datas[i]["index"])
            datas[i]["condition"] = tmp
        datalist.extend(datas)
        databaklist.extend(datasbak)
        print(len(datas))
        #记录各文件大小，用例保存时按照各文件进行保存
        filedict[filename] = (count, count+len(datas))
        count = count +len(datas)
    for data in datalist:
        uidlist.append(data["uid"])
    for i in range(len(uidlist)-1, -1, -1):
        #逆序遍历，避免访问越界
        #多轮交互为方便比较，仅保留首轮交互（首轮会带上后面的交互信息），并删除其他轮交互用例
        if (uidlist.index(uidlist[i]) != i):
            j = uidlist.index(uidlist[i])
            datalist[j]["delindex"].append(i)
            #将删除用例id记录在被保留的首轮交互中
            datalist[j]["condition"] = datalist[j]["condition"] + datalist[i]["condition"]
            del datalist[i]
    for data in datalist:
        conditionlist.append(data["condition"])
    multilist = []
    #多轮合并处理为单轮后，与单轮交互一并处理判断是否重复，若重复将用例记录下来
    for i in range(len(conditionlist)):
        if (conditionlist.index(conditionlist[i]) != i):
            multilist.append(i)
    delindexlist = []
    for multi in multilist:
        #根据重复用例回推在原始列表中重复的用例索引
        if (multiflag == "1"):
            delindexlist.extend(datalist[multi]["delindex"])
        elif (len(datalist[multi]["delindex"]) == 1):
            #多轮用例不进行去重
            delindexlist.extend(datalist[multi]["delindex"])
    for i in range(len(databaklist)):
        if i in delindexlist:
            databaklist[i].append(1)
        else:
            databaklist[i].append(None)
    count_src = len(databaklist)
    count_del = len(delindexlist)
    count_dest = count_src - count_del
    for key in filedict.keys():
        filepath = resultpath + "/" + key
        print(filedict)
        write_excel(filepath, databaklist[filedict[key][0]:filedict[key][1]],filetitledict[key], flag)
    return count_src, count_del, count_dest


def simplify(configdict):
    files = getFilelist(configdict["directory1"]["input"])
    conditions = configdict["rule1"]["condition"]
    flag = configdict["rule1"]["flag"]
    multiflag = configdict["rule1"]["multiflag"]
    nullflag = configdict["rule1"]["nullflag"]
    resultpath = configdict["directory1"]["result"]
    singlefileflag = configdict["rule1"]["singlefileflag"]
    count_src = 0
    count_del = 0
    count_dest = 0
    if (singlefileflag == "1"):
        for file in files:
            count_src_tmp, count_del_tmp, count_dest_tmp = de_weight([file], conditions, flag, multiflag, nullflag, resultpath)
            count_src = count_src + count_src_tmp
            count_del = count_del + count_del_tmp
            count_dest = count_dest + count_dest_tmp
    else:
        count_src,count_del,count_dest = de_weight(files, conditions, flag, multiflag, nullflag,resultpath)
    print("###################result######################")
    print("原始用例条数：",count_src)
    print("删除用例条数：", count_del)
    print("剩余用例条数：", count_dest)
    print("###################result######################")
    return 0

def casebackfill(configdict):
    inputpath = configdict["directory2"]["input"]
    inputfilelist = getFilelist(inputpath)
    outputpath = configdict["directory2"]["output"]
    outputfilelist = getFilelist(inputpath)
    key = configdict["rule2"]["key"]
    count = 0
    for inputfile in inputfilelist:
        outputfile = inputfile.replace(inputpath,outputpath)
        if os.path.exists(outputfile):
            wb = load_workbook(outputfile)
            ws = wb.active
            title_list = []
            top_flag = True
            for row in ws.rows:
                if top_flag:
                    for cell in row:
                        title_list.append(cell.value)
                    top_flag = False
            index = title_list.index(key)
            valuelist = []
            for cellObj in list(ws.columns)[index]:
                #print(cellObj.value)
                try:
                    jsobject = json.loads(cellObj.value)
                    value1 = jsobject.get("gram_variable_path")
                    value2 = jsobject.get("matched_regex")
                    if value1:
                        valuelist.append(value1)
                    elif value2:
                        valuelist.append(value2)
                    else:
                        valuelist.append(None)
                    count = count + 1
                except:
                    valuelist.append(None)
            valuelist[0] = "matched_regex"
            wbinput = load_workbook(inputfile)
            wsinput = wbinput.active
            max_column = wsinput.max_column
            for i in range(len(valuelist)):
                wsinput.cell(row=i+1, column=max_column+1, value=valuelist[i])
            wbinput.save(inputfile)
            wbinput.close()
        wb.close()
    print("###################result######################")
    print("含文法路径用例条数：",count)
    print("###################result######################")



def inspect(configdict):
    print('请检查config.ini参数function设置是否正确!')

def run():
    configdict = readconfig()
    function_dic = {
        '1' : simplify,
        '2' : casebackfill,
    }
    function_dic.get(configdict["function"]["function"], inspect)(configdict)

if __name__ == "__main__":
    run()